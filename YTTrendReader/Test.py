from io import BytesIO
import time
import os

import requests
import openpyxl as op
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pandas as pd

from googletrans import Translator
from collections import Counter
import nltk
from nltk.corpus import stopwords
import re

# 번역기 객체 생성
translator = Translator()

# 번역 및 키워드 추출 함수
def translate_and_extract_keywords(sentences, language_code):
    translated_sentences = []
    
    # 각 문장을 한국어로 번역
    for sentence in sentences:
        translated = translator.translate(sentence, src=language_code, dest='ko').text
        translated_sentences.append(translated)
    
    # 모든 번역된 문장을 하나로 합치기
    all_text = ' '.join(translated_sentences)
    
    # 불용어 제거 및 텍스트 전처리
    stop_words = set(stopwords.words('korean'))
    words = re.findall(r'\b\w+\b', all_text)
    words = [word for word in words if word not in stop_words]
    
    # 빈도 계산
    word_freq = Counter(words)
    
    return word_freq.most_common(10)  # 상위 10개의 키워드를 추출

# 각 국가별로 번역하고 키워드 추출하기


def get_youtube_trending_page_selenium(country_code='KR'):
    # ChromeDriver 경로 설정
    
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # 백그라운드 실행
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    # 브라우저 열기
    driver = webdriver.Chrome( options=options)
    
    # YouTube 인기 동영상 페이지 열기
    url = f"https://www.youtube.com/feed/trending?gl={country_code}"
    driver.get(url)
    time.sleep(1.5)
    # 페이지 로딩 대기
    
    last_height = driver.execute_script("return document.documentElement.scrollHeight")
    print(last_height)
    new_height = 0
    while new_height < last_height :
        new_height += 1000
        driver.execute_script(f"window.scrollTo(0, {new_height})")
        time.sleep(0.5)
        print(new_height)

    time.sleep(1.5)

    return driver

def parse_trending_videos_selenium(driver):
    videos = driver.find_elements(By.TAG_NAME, "ytd-video-renderer")
    
    video_data = []
    for index, video in enumerate(videos, start=1):
        try:
            # 제목, 채널, 조회수, URL 추출
            title_element = video.find_element(By.ID, 'video-title')
            title = title_element.text
            url = title_element.get_attribute('href')
            
            channel = video.find_element(By.XPATH, './/*[@id="text"]/a').text
            views = video.find_element(By.XPATH, './/*[@id="metadata-line"]/span[1]').text
            date = video.find_element(By.XPATH, './/*[@id="metadata-line"]/span[2]').text

            # 썸네일 URL 추출
            thumbnail_element = video.find_element(By.XPATH, './/*[@id="thumbnail"]/yt-image/img')
            thumbnail_url = thumbnail_element.get_attribute('src')
            print(thumbnail_element.get_attribute('src'))
            video_data.append({
                'Rank': index,
                'Title': title,
                'Channel': channel,
                'Views': views,
                'Date': date,
                'URL': url,
                'Thumbnail': thumbnail_url
            })
        except Exception as e:
            print(f"Error processing video {index}: {e}")
    
    return video_data


# Function to download and save thumbnail images, then embed them into Excel
def save_to_excel_with_images(video_data,code):
    wb = op.load_workbook('result.xlsx')
    ws = wb.create_sheet(code)

    # 헤더 추가
    headers = ['Rank', 'Title', 'Channel', 'Views','date', 'Thumbnail', 'URL']
    ws.append(headers)
    
    # 엑셀 셀 크기 설정 (썸네일 크기에 맞게)
    ws.column_dimensions['A'].width = 5  # 열 너비 설정
    ws.column_dimensions['B'].width = 54  # 열 너비 설정
    ws.column_dimensions['C'].width = 25  # 열 너비 설정
    ws.column_dimensions['D'].width = 16  # 열 너비 설정
    ws.column_dimensions['E'].width = 10  # 열 너비 설정
    ws.column_dimensions['F'].width = 16  # 열 너비 설정

    for index, video in enumerate(video_data, start=2):
        # 썸네일 이미지를 다운로드
        try:
            response = requests.get(video['Thumbnail'])
            img_data = BytesIO(response.content)
            img = PILImage.open(img_data)

            # 이미지 임시 저장 후 엑셀 삽입
            img_filename = f"thumbnail_{index}.png"
            img.save(img_filename)
            excel_img = Image(img_filename)
            excel_img.width, excel_img.height = (128, 72)
            ws.row_dimensions[index].height = 54  # 행 높이 설정
            
            # 썸네일을 셀에 삽입
            ws.add_image(excel_img, f'F{index}')
            
        except Exception as e:
            print(f"Failed to download image for video {video['Rank']}: {e}")
            

        # 텍스트 데이터 추가
        ws.append([video['Rank'], video['Title'], video['Channel'], video['Views'], video['Date'],'',video['URL']])

    # 엑셀 파일 저장
    
    wb.save('result.xlsx')
    print(f'Data saved.')

    for i in range(1, 100):
        file_name = f"thumbnail_{i}.png"  # 1.txt, 2.txt, ..., 99.txt와 같은 파일명 생성
        delete_file_if_exists(file_name)

def delete_file_if_exists(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} 파일이 삭제되었습니다.")
    else:
        print(f"{file_path} 파일이 존재하지 않습니다.")

# Main script
def main():
    #country_list = ['KR','US','VN','BR','ID','JP','TH','PH']
    country_list = ['US']
    wb = op.Workbook()
    wb.save('result.xlsx')

    for code in country_list:
        driver = get_youtube_trending_page_selenium(code)
        try:
            video_data = parse_trending_videos_selenium(driver)
            save_to_excel_with_images(video_data,code)
        finally:
            driver.quit()

    # nltk의 stopwords 다운로드
    nltk.download('stopwords')


    # 각국 문장 예시 (사용자 문장을 여기에 넣어주세요)
    sentences_by_country = {
        "US": ["Your Vietnamese sentences here..."],
        "JP": ["Your Japanese sentences here..."],
    }

    keywords_by_country = {}
    for country, sentences in sentences_by_country.items():
        if country == "US":
            language_code = 'en'
        elif country == "VN":
            language_code = 'vi'
        elif country == "BR":
            language_code = 'pt-BR'
        elif country == "ID":
            language_code = 'id'
        elif country == "JP":
            language_code = 'ja'
        elif country == "TH":
            language_code = 'th'
        elif country == "PH":
            language_code = 'fil'
        # 다른 국가도 마찬가지로 처리
        # ...
    
        keywords_by_country[country] = translate_and_extract_keywords(sentences, language_code)

    # 결과 출력
    for country, keywords in keywords_by_country.items():
        print(f"Country: {country}")
        print(f"Top Keywords: {keywords}")
        print("\n")


if __name__ == '__main__':
    main()
